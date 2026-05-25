import io
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Renk sabitleri ─────────────────────────────────────────────────────────────
_BG_DARK   = "0D1117"
_BG_SURF   = "161B22"
_BG_HEAD   = "1F2937"
_ACCENT    = "58A6FF"
_GREEN     = "3FB950"
_WARN      = "D29922"
_DANGER    = "F85149"
_ORANGE    = "FF7B00"
_TEXT      = "C9D1D9"
_MUTED     = "8B949E"

_VERDICT_MAP = {"accept": "KABUL", "conditional": "KOŞULLU", "reject": "RED"}
_RISK_MAP    = {"low": "Düşük", "medium": "Orta", "high": "Yüksek", "critical": "Kritik"}
_TYPE_MAP    = {"bump": "Çıkıntı", "hole": "Çukur", "missing": "Eksik Bölge"}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(color=_TEXT, bold=False, size=10) -> Font:
    return Font(color=color, bold=bold, size=size, name="Calibri")


def _border() -> Border:
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _header_row(ws, row: int, cols: list[str]) -> None:
    """Koyu arka planlı başlık satırı yazar."""
    for c, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill      = _fill(_BG_HEAD)
        cell.font      = _font(_ACCENT, bold=True, size=10)
        cell.alignment = _center()
        cell.border    = _border()


def _data_cell(ws, row: int, col: int, value, number_format=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _fill(_BG_SURF)
    cell.font      = _font()
    cell.alignment = _left()
    cell.border    = _border()
    if number_format:
        cell.number_format = number_format


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _verdict_color(v: str) -> str:
    return {"accept": _GREEN, "conditional": _WARN, "reject": _DANGER}.get(v, _MUTED)


def _risk_color(r: str) -> str:
    return {"low": _GREEN, "medium": _WARN, "high": _ORANGE, "critical": _DANGER}.get(r, _MUTED)


# ── Sayfa 1: Analiz Özeti ──────────────────────────────────────────────────────
def _sheet_summary(wb: Workbook, item: dict) -> None:
    ws = wb.create_sheet("Analiz Özeti")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    stats = item["result"]["stats"]
    ts    = datetime.fromisoformat(item["timestamp"]).strftime("%Y-%m-%d %H:%M")

    # Başlık
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value     = f"LIFT UP — Analiz Özeti  |  #{item['id']}  |  {item['model']}  |  {ts}"
    title.fill      = _fill(_BG_DARK)
    title.font      = _font(_ACCENT, bold=True, size=12)
    title.alignment = _center()
    ws.row_dimensions[1].height = 26

    _header_row(ws, 2, ["Parametre", "Değer", "Birim", "Açıklama"])
    ws.row_dimensions[2].height = 20

    rows = [
        ("Model",           item["model"],                          "",    "Referans STL dosyası"),
        ("Analiz ID",       item["id"],                             "",    "Veritabanı kaydı"),
        ("Tarih",           ts,                                     "",    "Analiz zamanı"),
        ("Nokta Sayısı",    item["n_points"],                       "adet","Tarama yoğunluğu"),
        ("Sensör Gürültüsü",item["noise_std"],                      "mm",  "Gaussian σ"),
        ("Hata Sayısı",     item["defect_count"],                   "adet","Simüle edilen hata"),
        ("",                "",                                     "",    ""),
        ("Uyum Oranı",      stats["conformance"],                   "%",   "Tolerans içindeki nokta yüzdesi"),
        ("Tolerans",        stats["tolerance_mm"],                  "mm",  "3-sigma eşiği"),
        ("Karar",           _VERDICT_MAP.get(stats["verdict"], "?"),"",    "Kalite kararı"),
        ("Genel Risk",      _RISK_MAP.get(stats["overall_risk"],""),"",    "En yüksek hata riski"),
        ("",                "",                                     "",    ""),
        ("RMS Sapma",       stats["rms"],                           "mm",  "Genel hata büyüklüğü"),
        ("Maks Sapma",      stats["max_deviation"],                 "mm",  "En kötü nokta"),
        ("Ort. Sapma",      stats["mean_deviation"],                "mm",  "Ortalama hata"),
        ("ICP RMSE",        stats["icp_rmse"],                      "mm",  "Hizalama hatası"),
        ("Anomali Sayısı",  stats["anomaly_count"],                 "nokta","Tolerans dışı"),
        ("DBSCAN Küme",     stats.get("anomaly_clusters", "—"),     "adet","Anomali küme sayısı"),
        ("Toplam Nokta",    stats["total_points"],                  "adet","Tarama noktası"),
        ("",                "",                                     "",    ""),
        ("ICP Yöntemi",     stats.get("icp_method", "—"),           "",    "Hizalama algoritması"),
        ("Sapma Yöntemi",   stats.get("deviation_method", "—"),     "",    "Mesafe hesabı"),
    ]

    for r, (param, val, unit, desc) in enumerate(rows, 3):
        ws.row_dimensions[r].height = 18
        if param == "":
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = _fill(_BG_DARK)
            continue
        _data_cell(ws, r, 1, param)
        ws.cell(row=r, column=1).font = _font(_MUTED)

        cell_val = ws.cell(row=r, column=2, value=val)
        cell_val.fill      = _fill(_BG_SURF)
        cell_val.alignment = _left()
        cell_val.border    = _border()
        # Sayısal değerlere renk
        if param in ("Uyum Oranı", "Karar", "Genel Risk"):
            if param == "Karar":
                cell_val.font = Font(color=_verdict_color(stats["verdict"]),
                                     bold=True, size=10, name="Calibri")
            elif param == "Genel Risk":
                cell_val.font = Font(color=_risk_color(stats["overall_risk"]),
                                     bold=True, size=10, name="Calibri")
            else:
                conf = stats["conformance"]
                c = _GREEN if conf >= 95 else _WARN if conf >= 85 else _ORANGE if conf >= 70 else _DANGER
                cell_val.font = Font(color=c, bold=True, size=10, name="Calibri")
        else:
            cell_val.font = _font(bold=True)

        _data_cell(ws, r, 3, unit)
        _data_cell(ws, r, 4, desc)
        ws.cell(row=r, column=4).font = _font(_MUTED)

    _set_col_widths(ws, [22, 28, 8, 42])


# ── Sayfa 2: Hata Bölgeleri ───────────────────────────────────────────────────
def _sheet_defects(wb: Workbook, item: dict) -> None:
    ws = wb.create_sheet("Hata Bölgeleri")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    defects = item["result"].get("defect_regions", [])

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value     = f"Hata Bölgeleri  |  Analiz #{item['id']}  |  {item['model']}"
    title.fill      = _fill(_BG_DARK)
    title.font      = _font(_ACCENT, bold=True, size=12)
    title.alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["#", "Tip", "Büyüklük (mm)", "Yarıçap (mm)",
               "Risk Seviyesi", "Merkez X", "Merkez Y", "Merkez Z"]
    _header_row(ws, 2, headers)

    if not defects:
        ws.merge_cells("A3:H3")
        c = ws["A3"]
        c.value     = "Bu analizde hata bölgesi bulunmamaktadır."
        c.fill      = _fill(_BG_SURF)
        c.font      = _font(_MUTED)
        c.alignment = _center()
    else:
        for r, d in enumerate(defects, 3):
            ws.row_dimensions[r].height = 18
            risk_level = d.get("risk", {}).get("level", "high")
            rc = _risk_color(risk_level)
            center = d.get("center", [0, 0, 0])

            vals = [
                r - 2,
                _TYPE_MAP.get(d["type"], d["type"]),
                round(d["magnitude"], 3),
                round(d["radius"], 2),
                _RISK_MAP.get(risk_level, risk_level),
                round(center[0], 3) if len(center) > 0 else "—",
                round(center[1], 3) if len(center) > 1 else "—",
                round(center[2], 3) if len(center) > 2 else "—",
            ]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.fill      = _fill(_BG_SURF)
                cell.alignment = _left()
                cell.border    = _border()
                if c_idx == 5:  # Risk sütunu renkli
                    cell.font = Font(color=rc, bold=True, size=10, name="Calibri")
                else:
                    cell.font = _font()

    _set_col_widths(ws, [5, 16, 16, 14, 14, 13, 13, 13])


# ── Sayfa 3: Sapma Verileri ───────────────────────────────────────────────────
def _sheet_distances(wb: Workbook, item: dict, max_rows: int = 2000) -> None:
    ws = wb.create_sheet("Sapma Verileri")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    dists    = np.array(item["result"].get("distances", []))
    pts      = np.array(item["result"].get("points", []))
    anom     = item["result"].get("anomalies", [])
    tol      = item["result"]["stats"]["tolerance_mm"]
    n        = min(len(dists), max_rows)

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value     = (f"Sapma Verileri  |  İlk {n} nokta gösterilmektedir"
                       f"  |  Toplam: {len(dists):,}")
    title.fill      = _fill(_BG_DARK)
    title.font      = _font(_ACCENT, bold=True, size=12)
    title.alignment = _center()
    ws.row_dimensions[1].height = 26

    _header_row(ws, 2, ["#", "X (mm)", "Y (mm)", "Z (mm)", "Sapma (mm)", "Durum"])

    for r in range(n):
        row_idx = r + 3
        ws.row_dimensions[row_idx].height = 16
        dist  = float(dists[r])
        is_ok = dist < tol
        status = "OK" if is_ok else "ANOMALİ"
        sc     = _GREEN if is_ok else _DANGER

        x = round(float(pts[r][0]), 3) if r < len(pts) else "—"
        y = round(float(pts[r][1]), 3) if r < len(pts) else "—"
        z = round(float(pts[r][2]), 3) if r < len(pts) else "—"

        for c_idx, val in enumerate([r + 1, x, y, z, round(dist, 4), status], 1):
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.fill      = _fill(_BG_SURF)
            cell.alignment = _left()
            cell.border    = _border()
            if c_idx == 6:
                cell.font = Font(color=sc, bold=True, size=9, name="Calibri")
            else:
                cell.font = _font(size=9)

    _set_col_widths(ws, [8, 13, 13, 13, 14, 12])


# ── Ana fonksiyon ──────────────────────────────────────────────────────────────
def generate_excel(item: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)   # varsayılan boş sayfayı kaldır

    _sheet_summary(wb, item)
    _sheet_defects(wb, item)
    _sheet_distances(wb, item)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_history_excel(history: list) -> bytes:
    """Tüm geçmişi tek sayfada özet olarak dışa aktarır."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analiz Geçmişi"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value     = f"LIFT UP — Analiz Geçmişi  |  {len(history)} kayıt"
    title.fill      = _fill(_BG_DARK)
    title.font      = _font(_ACCENT, bold=True, size=12)
    title.alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["ID", "Tarih", "Model", "Nokta", "σ (mm)", "Hata",
               "Uyum %", "RMS (mm)", "Maks (mm)", "Karar", "Risk"]
    _header_row(ws, 2, headers)

    for r, h in enumerate(history, 3):
        ws.row_dimensions[r].height = 18
        verdict = h.get("verdict", "")
        risk    = h.get("overall_risk", "")
        ts      = datetime.fromisoformat(h["timestamp"]).strftime("%Y-%m-%d %H:%M")
        vals    = [
            h["id"], ts, h["model"], h["n_points"],
            h["noise_std"], h["defect_count"],
            h["conformance"], h["rms"], h["max_deviation"],
            _VERDICT_MAP.get(verdict, verdict),
            _RISK_MAP.get(risk, risk),
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.fill      = _fill(_BG_SURF)
            cell.alignment = _left()
            cell.border    = _border()
            if c_idx == 10:
                cell.font = Font(color=_verdict_color(verdict),
                                 bold=True, size=10, name="Calibri")
            elif c_idx == 11:
                cell.font = Font(color=_risk_color(risk),
                                 bold=True, size=10, name="Calibri")
            else:
                cell.font = _font()

    _set_col_widths(ws, [6, 17, 16, 8, 8, 7, 9, 10, 10, 11, 10])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
